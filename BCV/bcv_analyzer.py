import contextlib
import csv
import itertools
import json
import os
import sys
import threading
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Annotated, Any, Callable
from urllib.parse import urlparse

import questionary
import typer
from rich.console import Console
from rich.panel import Panel

ALLOWED_TABLES = ("request", "ad", "slot", "candidate", "auction", "ack")
SRC_CATALOG = "mrm_log_flat"
SRC_SCHEMA = "default"
BCV_CATALOG = "etl"
BCV_SCHEMA = "public_test1"
EXCLUDED_DESCRIBE_COLUMNS = {"extra", "comment"}
UI_CONSOLE = Console()
_SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = _SCRIPT_DIR / "output"
FIELD_SIZE_DIR = _SCRIPT_DIR / "field_size"
ETL_FIELDS_PATH = _SCRIPT_DIR / "etl_fields.json"
SOS_FIELDS_PATH = _SCRIPT_DIR / "sos_fields.csv"
EXCLUDE_FIELDS_PATH = _SCRIPT_DIR / "exclude.csv"
APP_VERSION = "v0.1"
USAGE_COLUMNS = (
    "usage:ETL",
    "usage:SOS",
    "usage:Insights",
    "usage:Arena",
    "usage:LQS",
    "usage:CP",
    "usage:AF",
    "usage:Others",
)
USAGE_QUERY_BATCH_SIZE = 500
USAGE_QUERY_MAX_RETRIES = 3
VALUE_VALIDATION_BATCH_SIZE = 500
VALUE_VALIDATION_REQUEST_TIMEOUT = 60  # seconds — validation queries can be slow
# Key columns used to join SRC and BCV rows, configured per table.
# For tables not listed here, value validation is skipped.
TABLE_KEY_COLUMNS: dict[str, list[str]] = {
    "request": ["request__transaction_id"],
    "slot": ["request__transaction_id", "slot__index"],
    "ad": ["request__transaction_id", "advertisement__ad_id", "advertisement__ad_replica_id"],
    "ack": ["request__transaction_id", "ack__kafka_msg_key", "ack__ack_entity_type"],
    "candidate": ["request__transaction_id", "candidate__ctx_index"],
    "auction": ["request__transaction_id", "auction__ctx_index"],
}
# Columns whose SQL literal must be unquoted (int). All others default to quoted string.
TABLE_KEY_COLUMN_TYPES: dict[str, str] = {
    "slot__index": "int",
    "advertisement__ad_id": "int",
    "advertisement__ad_replica_id": "int",
    "candidate__ctx_index": "int",
    "auction__ctx_index": "int",
}


class RunMode(str, Enum):
    FULL_RUN = "Full Run"
    VALIDATION_ONLY = "Validation Only"


APP_BANNER = r"""
 ____   ____ __     __     _                _
| __ ) / ___|\ \   / /    / \   _ __   ___ | |_   _ _______ _ __
|  _ \| |     \ \ / /    / _ \ | '_ \ / _ \| | | | |_  / _ \ '__|
| |_) | |___   \ V /    / ___ \| | | | (_| | | |_| |/ /  __/ |
|____/ \____|   \_/    /_/   \_\_| |_|\__,_|_|\__, /___\___|_|
                                              |___/
"""


def build_http_headers(auth_token: str | None, auth_header: str | None) -> dict[str, str]:
    if not auth_token:
        return {}

    header_name = auth_header or "Authorization"
    if header_name.lower() == "authorization":
        return {header_name: f"Bearer {auth_token}"}
    return {header_name: auth_token}


def current_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log_info(message: str) -> None:
    print(f"[{current_timestamp()}] {message}")


def log_progress(message: str, done: bool = False) -> None:
    print(f"\r[{current_timestamp()}] {message}", end="\n" if done else "", flush=True)


@contextlib.contextmanager
def spinning_cursor(message: str = "Executing Presto query"):
    """Display a spinning cursor animation while the enclosed block runs."""
    stop_event = threading.Event()
    frames = itertools.cycle(["|", "/", "-", "\\"])

    def _spin() -> None:
        while not stop_event.is_set():
            print(f"\r[{current_timestamp()}] {message} {next(frames)}", end="", flush=True)
            stop_event.wait(timeout=1.0)
        # Clear the spinner line
        print(f"\r{' ' * (len(message) + 35)}\r", end="", flush=True)

    thread = threading.Thread(target=_spin, daemon=True)
    thread.start()
    try:
        yield
    finally:
        stop_event.set()
        thread.join()


def build_http_session(headers: dict[str, str]):
    import requests

    session = requests.Session()
    session.headers.update(headers)
    return session


def normalize_connection_kwargs(connection_kwargs: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(connection_kwargs)
    normalized["http_scheme"] = "https"
    normalized.setdefault("port", 8080)

    host = normalized.get("host")
    if not host:
        return normalized

    raw_host = str(host).strip()
    parsed = urlparse(raw_host if "://" in raw_host else f"https://{raw_host}")
    if parsed.hostname:
        normalized["host"] = parsed.hostname
    if parsed.port is not None:
        normalized["port"] = parsed.port

    return normalized


def build_engine_url(host: str, port: int, user: str, catalog: str, schema: str) -> str:
    return f"trino://{user}@{host}:{port}/{catalog}/{schema}"


def build_full_table_name(table: str) -> str:
    return f"{SRC_CATALOG}.{SRC_SCHEMA}.{table}"


def build_describe_sql(table: str) -> str:
    return f"DESCRIBE {build_full_table_name(table)}"


def build_bcv_table_name(table: str) -> str:
    return f"{BCV_CATALOG}.{BCV_SCHEMA}.{table}"


def build_bcv_describe_sql(table: str) -> str:
    return f"DESCRIBE {build_bcv_table_name(table)}"


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def build_usage_sql(table: str, column: str) -> str:
    return build_usage_sql_batch(table, [column])


def build_usage_sql_batch(table: str, columns: list[str]) -> str:
    col_list = ", ".join(sql_literal(col) for col in columns)
    return f"""
SELECT
    a.col AS column_name,
    CASE
        WHEN a.user IN ('sa-dataapp-insights', 'sa-dmo-aqs', 'sa-dataapp-yield', 'svc-ciec-sct', 'sa-trust_standards', 'sa-analytics-scrum') THEN 'Insights'
        WHEN a.user IN ('publisher') THEN 'CP'
        WHEN a.user IN ('sa-presto-af-etl') THEN 'AF'
        WHEN a.source LIKE '%Arena%' THEN 'Arena'
        WHEN a.source LIKE '%lqs%' THEN 'LQS'
        ELSE 'Others'
    END AS user,
    count(*) as usage_count
FROM (
    SELECT DISTINCT
        q.pid,
        q.user,
        q.source,
        q.create_date,
        pcu.catalog,
        pcu.schema,
        pcu.table_name,
        p.col
    FROM etl.pgw_etl."query" q
    JOIN etl.pgw_etl."presto_column_usage" pcu on pcu.pid = q.pid
    CROSS JOIN UNNEST(pcu.columns) AS p(col)
    WHERE q.create_date >= DATE '2026-01-01'
    AND pcu.create_date >= DATE '2026-01-01'
    AND pcu.catalog = {sql_literal(SRC_CATALOG)}
    AND pcu.schema = {sql_literal(SRC_SCHEMA)}
    AND pcu.table_name = {sql_literal(table)}
    AND env = 'prd'
    AND q.user NOT IN ('sqyang', 'yjgou', 'kbhargava', 'yuwang', 'zhfan')
    -- AND q.source NOT IN ('presto-python-client')
    AND p.col IN ({col_list})
    AND (NOT regexp_like(q.query, '(?i)select\\s*\\*'))
    AND error_type is null AND error_name is null
) a
GROUP BY 1, 2
ORDER BY 1, 3 DESC
""".strip()


def build_connect_args(connection_kwargs: dict[str, Any]) -> dict[str, Any]:
    headers = connection_kwargs.get("http_headers") or {}
    return {
        "http_scheme": connection_kwargs.get("http_scheme", "https"),
        "http_session": build_http_session(headers),
        "request_timeout": connection_kwargs.get("request_timeout", 5.0),
        "max_attempts": 1,
    }


def format_rows(
    rows: list[tuple[Any, ...]], description: list[tuple[Any, ...]]
) -> list[dict[str, Any]]:
    columns = [column[0] for column in description]
    return [dict(zip(columns, row)) for row in rows]


def remove_describe_metadata_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            name: value
            for name, value in row.items()
            if name.lower() not in EXCLUDED_DESCRIBE_COLUMNS
        }
        for row in rows
    ]


def get_case_insensitive(row: dict[str, Any], key: str) -> Any:
    for name, value in row.items():
        if name.lower() == key.lower():
            return value
    return ""


def normalize_field_name_for_size_lookup(field_name: str) -> str:
    return field_name.replace("__", ".")


def lookup_size(src_field: str, field_sizes: dict[str, float]) -> float | str:
    if not src_field:
        return ""

    size = field_sizes.get(normalize_field_name_for_size_lookup(src_field))
    if size is None:
        return ""
    return round(size, 2)


def load_field_sizes(table: str, field_size_dir: Path = FIELD_SIZE_DIR) -> dict[str, float]:
    from openpyxl import load_workbook

    workbook_path = field_size_dir / f"{table}_raw_size_in_TiB.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    field_index = headers.index("Field Name")
    size_index = headers.index("Size (TiB)")

    field_sizes = {}
    for row in rows:
        field_name = row[field_index]
        size = row[size_index]
        if field_name and size is not None:
            field_sizes[str(field_name)] = float(size)
    return field_sizes


def load_etl_fields(table: str, etl_fields_path: Path = ETL_FIELDS_PATH) -> set[str]:
    if not etl_fields_path.exists():
        return set()

    data = json.loads(etl_fields_path.read_text(encoding="utf-8"))
    return {str(field) for field in data.get(table, [])}


def add_etl_usage_info(
    table: str,
    rows: list[dict[str, Any]],
    etl_fields_path: Path = ETL_FIELDS_PATH,
) -> list[dict[str, Any]]:
    etl_fields = load_etl_fields(table, etl_fields_path)
    for row in rows:
        row.setdefault("usage:ETL", "")
        if not is_missing_bcv_column(row):
            continue
        normalized_src_field = normalize_field_name_for_size_lookup(str(row.get("src_field") or ""))
        if normalized_src_field in etl_fields:
            row["usage:ETL"] = "Y"
    return rows


def load_sos_fields(table: str, sos_fields_path: Path = SOS_FIELDS_PATH) -> set[str]:
    if not sos_fields_path.exists():
        return set()
    fields: set[str] = set()
    with sos_fields_path.open(encoding="utf-8", newline="") as f:
        for sos_row in csv.DictReader(f):
            tbl = (sos_row.get("table") or "").strip()
            field = (sos_row.get("column") or "").strip()
            if tbl == table and field:
                fields.add(field)
    return fields


def add_sos_usage_info(
    table: str,
    rows: list[dict[str, Any]],
    sos_fields_path: Path = SOS_FIELDS_PATH,
) -> list[dict[str, Any]]:
    sos_fields = load_sos_fields(table, sos_fields_path)
    for row in rows:
        row.setdefault("usage:SOS", "")
        if not is_missing_bcv_column(row):
            continue
        normalized_src_field = normalize_field_name_for_size_lookup(str(row.get("src_field") or ""))
        if normalized_src_field in sos_fields:
            row["usage:SOS"] = "Y"
    return rows


def load_excluded_fields(table: str, exclude_path: Path = EXCLUDE_FIELDS_PATH) -> set[str]:
    """Return the set of src_field names (__ notation) to exclude for *table*."""
    if not exclude_path.exists():
        return set()
    fields: set[str] = set()
    with exclude_path.open(encoding="utf-8", newline="") as f:
        for ex_row in csv.DictReader(f):
            tbl = (ex_row.get("table") or "").strip()
            col = (ex_row.get("column") or "").strip()
            if tbl == table and col:
                fields.add(col)
    return fields


def compare_schema_rows(
    table: str,
    src_rows: list[dict[str, Any]],
    bcv_rows: list[dict[str, Any]],
    field_sizes: dict[str, float] | None = None,
) -> list[dict[str, Any]]:
    field_sizes = field_sizes or {}
    bcv_by_name = {get_case_insensitive(row, "column"): row for row in bcv_rows}
    compared = []

    for src_row in src_rows:
        src_name = get_case_insensitive(src_row, "column")
        src_type = get_case_insensitive(src_row, "type")
        bcv_row = bcv_by_name.get(src_name)
        bcv_type = get_case_insensitive(bcv_row, "type") if bcv_row else ""
        if bcv_row and src_type == bcv_type:
            status = "MATCHED"
        elif bcv_row:
            status = "MATCHED - TYPE DIFF"
        else:
            status = "DIFF"
        compared.append(
            {
                "status": status,
                "src_field": src_name,
                "src_type": src_type,
                "bcv_field": get_case_insensitive(bcv_row, "column") if bcv_row else "",
                "bcv_type": bcv_type,
                "size": lookup_size(src_name, field_sizes),
            }
        )

    src_names = {get_case_insensitive(row, "column") for row in src_rows}
    for bcv_row in bcv_rows:
        bcv_name = get_case_insensitive(bcv_row, "column")
        if bcv_name in src_names:
            continue
        compared.append(
            {
                "status": "DIFF",
                "src_field": "",
                "src_type": "",
                "bcv_field": bcv_name,
                "bcv_type": get_case_insensitive(bcv_row, "type"),
                "size": "",
            }
        )

    return compared


def count_missing_bcv_columns(rows: list[dict[str, Any]]) -> int:
    return sum(
        1
        for row in rows
        if row.get("status") == "DIFF" and not row.get("bcv_field") and not row.get("bcv_type")
    )


def is_missing_bcv_column(row: dict[str, Any]) -> bool:
    return (
        row.get("status") == "DIFF"
        and bool(row.get("src_field"))
        and not row.get("bcv_field")
        and not row.get("bcv_type")
    )


def is_type_mismatch(row: dict[str, Any]) -> bool:
    """True when the field exists in both SRC and BCV but the types differ."""
    return row.get("status") == "MATCHED - TYPE DIFF"


def add_usage_info(
    table: str,
    rows: list[dict[str, Any]],
    connection_kwargs: dict[str, Any],
) -> list[dict[str, Any]]:
    for row in rows:
        for column in USAGE_COLUMNS:
            row.setdefault(column, "")

    missing_rows = [row for row in rows if is_missing_bcv_column(row)]
    total = len(missing_rows)
    if total == 0:
        return rows

    row_by_field: dict[str, dict[str, Any]] = {str(r["src_field"]): r for r in missing_rows}
    num_batches = (total + USAGE_QUERY_BATCH_SIZE - 1) // USAGE_QUERY_BATCH_SIZE

    for batch_idx in range(num_batches):
        batch_start = batch_idx * USAGE_QUERY_BATCH_SIZE
        batch_end = min(batch_start + USAGE_QUERY_BATCH_SIZE, total)
        batch_columns = [str(r["src_field"]) for r in missing_rows[batch_start:batch_end]]

        usage_rows: list[dict[str, Any]] = []
        for attempt in range(USAGE_QUERY_MAX_RETRIES + 1):
            try:
                with spinning_cursor(
                    f"Retrieving usage info for missing columns in BCV:"
                    f" batch {batch_idx + 1}/{num_batches} (columns {batch_start + 1}–{batch_end} of {total})"
                ):
                    usage_rows = execute_sql(
                        build_usage_sql_batch(table, batch_columns),
                        connection_kwargs=connection_kwargs,
                    )
                break
            except Exception as exc:
                if attempt < USAGE_QUERY_MAX_RETRIES:
                    log_info(
                        f"Retry {attempt + 1}/{USAGE_QUERY_MAX_RETRIES} for batch"
                        f" {batch_idx + 1}/{num_batches}: {exc}"
                    )
                else:
                    log_info(
                        f"Failed to retrieve usage info for batch {batch_idx + 1}/{num_batches}"
                        f" after {USAGE_QUERY_MAX_RETRIES} retries: {exc}"
                    )

        for usage_row in usage_rows:
            col_name = str(get_case_insensitive(usage_row, "column_name"))
            user = get_case_insensitive(usage_row, "user")
            usage_column = f"usage:{user}"
            if col_name in row_by_field and usage_column in USAGE_COLUMNS:
                row_by_field[col_name][usage_column] = get_case_insensitive(
                    usage_row, "usage_count"
                )

    return rows, missing_rows


def execute_sql(
    sql: str,
    connection_kwargs: dict[str, Any],
    engine_factory: Callable[..., Any] | None = None,
) -> list[dict[str, Any]]:
    kwargs = normalize_connection_kwargs(connection_kwargs)
    if engine_factory is None:
        from sqlalchemy import create_engine

        engine_factory = create_engine

    engine = engine_factory(
        build_engine_url(
            host=kwargs["host"],
            port=kwargs["port"],
            user=kwargs["user"],
            catalog=kwargs["catalog"],
            schema=kwargs["schema"],
        ),
        connect_args=build_connect_args(kwargs),
    )

    connection = engine.raw_connection()
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        return format_rows([tuple(row) for row in rows], cursor.description)
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception as exc:
                print(f"Warning: failed to close cursor cleanly: {exc}", file=sys.stderr)
        try:
            connection.close()
        except Exception as exc:
            print(f"Warning: failed to close connection cleanly: {exc}", file=sys.stderr)


def print_rows_as_json(rows: list[dict[str, Any]]) -> None:
    print(json.dumps(rows, ensure_ascii=False, indent=2, default=str))


def write_rows_as_json_file(rows: list[dict[str, Any]], table: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / f"{table}.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def write_rows_as_csv_file(
    rows: list[dict[str, Any]], filename: str, fieldnames: tuple[str, ...]
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUTPUT_DIR / filename).open("w", encoding="utf-8", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_value_validation_batch_id(now: datetime | None = None) -> str:
    batch_time = (now or datetime.now()) - timedelta(hours=24)
    return batch_time.replace(minute=0, second=0, microsecond=0).strftime("%Y%m%d%H%M%S")


def load_matched_columns_from_result_csv(result_path: Path) -> list[str]:
    with result_path.open(encoding="utf-8", newline="") as result_file:
        return [
            row["src_field"]
            for row in csv.DictReader(result_file)
            if row.get("status") in ("MATCHED", "MATCHED - TYPE DIFF") and row.get("src_field")
        ]


# src_type values that indicate a nested array-of-string parent structure node
PARENT_STRUCTURE_TYPES: frozenset[str] = frozenset(
    {
        "varchar",
        "array(varchar)",
        "array(array(varchar))",
        "array(array(array(varchar)))",
    }
)


def identify_parent_structure_nodes(result_path: Path) -> dict[str, str]:
    """Return a mapping of src_field → src_type for parent structure nodes.

    Both conditions must hold:
    1. src_type is one of PARENT_STRUCTURE_TYPES
    2. At least one child column ({src_field}__*) exists anywhere in the CSV
    """
    with result_path.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    all_src_fields = {row.get("src_field", "") for row in rows if row.get("src_field")}

    parent_nodes: dict[str, str] = {}
    for row in rows:
        src_field = row.get("src_field", "")
        src_type = (row.get("src_type") or "").lower().strip()
        if not src_field or src_type not in PARENT_STRUCTURE_TYPES:
            continue
        if any(f.startswith(f"{src_field}__") for f in all_src_fields):
            parent_nodes[src_field] = src_type
    return parent_nodes


def get_table_key_columns(table: str) -> list[str]:
    return TABLE_KEY_COLUMNS.get(table, [])


def extract_row_key(row: dict[str, Any], key_columns: list[str]) -> tuple:
    return tuple(get_case_insensitive(row, col) for col in key_columns)


def format_row_key(key: tuple) -> str:
    return " / ".join(str(v) for v in key)


def build_key_in_clause(key_columns: list[str], keys: list[tuple]) -> str:
    """Build a SQL IN predicate for one or more key columns."""

    def _fmt(col: str, value: Any) -> str:
        if TABLE_KEY_COLUMN_TYPES.get(col) == "int":
            return str(int(value))
        return sql_literal(str(value))

    if len(key_columns) == 1:
        col = key_columns[0]
        key_list = ",\n".join(f"      {_fmt(col, k[0])}" for k in keys)
        return f"{col} IN (\n{key_list}\n  )"
    col_tuple = "(" + ", ".join(key_columns) + ")"
    key_rows = ",\n".join(
        "      (" + ", ".join(_fmt(col, v) for col, v in zip(key_columns, k)) + ")" for k in keys
    )
    return f"{col_tuple} IN (\n{key_rows}\n  )"


def get_row_keys(rows: list[dict[str, Any]], key_columns: list[str]) -> list[tuple]:
    keys = []
    for row in rows:
        key = extract_row_key(row, key_columns)
        if all(v is not None and v != "" for v in key):
            keys.append(key)
    return keys


def build_value_validation_sql_batches(
    table: str,
    columns: list[str],
    batch_id: str,
    key_columns: list[str],
    batch_size: int = VALUE_VALIDATION_BATCH_SIZE,
    limit: int = 10,
) -> list[dict[str, Any]]:
    sql_batches = []
    for batch_start in range(0, len(columns), batch_size):
        batch_columns = columns[batch_start : batch_start + batch_size]
        # Always include all key columns so every batch can be joined
        missing_keys = [k for k in key_columns if k not in batch_columns]
        if missing_keys:
            batch_columns = missing_keys + batch_columns
        select_list = ",\n".join(f"    {column}" for column in batch_columns)
        src_sql = (
            '-- pgw_tags: {"client_tags":"2XLARGE"}\n'
            "SELECT\n"
            f"{select_list}\n"
            f"FROM {build_full_table_name(table)} TABLESAMPLE BERNOULLI (1)\n"
            f"WHERE bitwise_and(request__bit_flags, 576460752303423488) > 0\n"
            f"  AND process_batch_id = {sql_literal(batch_id)}\n"
            f"LIMIT {limit}"
        )
        sql_batches.append({"src_sql": src_sql, "columns": batch_columns})
    return sql_batches


def build_value_validation_bcv_sql(
    table: str,
    columns: list[str],
    batch_id: str,
    key_columns: list[str],
    keys: list[tuple],
    limit: int = 10,
) -> str:
    select_list = ",\n".join(f"    {column}" for column in columns)
    key_clause = build_key_in_clause(key_columns, keys)
    return (
        '-- pgw_tags: {"client_tags":"2XLARGE"}\n'
        "SELECT\n"
        f"{select_list}\n"
        f"FROM {build_bcv_table_name(table)}\n"
        f"WHERE process_batch_id = {sql_literal(batch_id)}\n"
        f"  AND {key_clause}\n"
        f"LIMIT {limit}"
    )


def build_value_validation_src_sql_by_keys(
    table: str,
    columns: list[str],
    batch_id: str,
    key_columns: list[str],
    keys: list[tuple],
) -> str:
    """SRC SQL for batch 2+: target the same rows sampled in batch 1."""
    select_list = ",\n".join(f"    {column}" for column in columns)
    key_clause = build_key_in_clause(key_columns, keys)
    return (
        '-- pgw_tags: {"client_tags":"2XLARGE"}\n'
        "SELECT\n"
        f"{select_list}\n"
        f"FROM {build_full_table_name(table)}\n"
        f"WHERE process_batch_id = {sql_literal(batch_id)}\n"
        f"  AND {key_clause}\n"
    )


def compare_value_validation_results(
    src_rows: list[dict[str, Any]],
    bcv_rows: list[dict[str, Any]],
    columns: list[str],
    key_columns: list[str],
    use_global_equiv: bool = False,
) -> dict[str, Any]:
    key_set = set(key_columns)
    value_columns = [column for column in columns if column not in key_set]
    bcv_by_key: dict[tuple, dict[str, Any]] = {
        extract_row_key(row, key_columns): row for row in bcv_rows
    }
    # field_diffs: field → list of {key, src, bcv} for actual mismatches
    field_diffs: dict[str, list[dict[str, Any]]] = {}
    # globally_equiv_diffs: diffs that are semantically equivalent (pass in equiv mode)
    globally_equiv_diffs: dict[str, list[dict[str, Any]]] = {}
    matched_transaction_count = 0

    for src_row in src_rows:
        key = extract_row_key(src_row, key_columns)
        key_display = format_row_key(key)
        bcv_row = bcv_by_key.get(key)
        if bcv_row is None:
            continue
        matched_transaction_count += 1
        for column in value_columns:
            src_value = get_case_insensitive(src_row, column)
            bcv_value = get_case_insensitive(bcv_row or {}, column)
            if not _null_safe_equal(src_value, bcv_value):
                diff_entry = {"key": key_display, "src": src_value, "bcv": bcv_value}
                if use_global_equiv and _globally_equivalent(src_value, bcv_value):
                    globally_equiv_diffs.setdefault(column, []).append(diff_entry)
                else:
                    field_diffs.setdefault(column, []).append(diff_entry)

    # preserve original column order
    # globally_equiv_fields: no real diffs, but some equiv diffs present
    globally_equiv_fields = [
        c for c in value_columns if c not in field_diffs and c in globally_equiv_diffs
    ]
    matched_fields = [
        c for c in value_columns if c not in field_diffs and c not in globally_equiv_diffs
    ]
    mismatched_field_list = [c for c in value_columns if c in field_diffs]
    total_field_count = len(value_columns)
    matched_field_count = len(matched_fields)
    globally_equiv_field_count = len(globally_equiv_fields)
    mismatched_field_count = len(mismatched_field_list)
    effective_matched = matched_field_count + globally_equiv_field_count
    return {
        "total_transaction_count": len(src_rows),
        "matched_transaction_count": matched_transaction_count,
        "matched_fields": matched_fields,
        "globally_equiv_fields": globally_equiv_fields,
        "globally_equiv_diffs": globally_equiv_diffs,
        "globally_equiv_field_count": globally_equiv_field_count,
        "mismatched_fields": mismatched_field_list,
        "field_diffs": field_diffs,
        "total_field_count": total_field_count,
        "matched_field_count": matched_field_count,
        "mismatched_field_count": mismatched_field_count,
        "matched_field_ratio": (effective_matched / total_field_count * 100)
        if total_field_count
        else 0,
        "mismatched_field_ratio": (mismatched_field_count / total_field_count * 100)
        if total_field_count
        else 0,
    }


def print_value_validation_summary(summary: dict[str, Any]) -> None:
    matched_tx = summary["matched_transaction_count"]
    total_tx = summary["total_transaction_count"]
    matched_f = summary["matched_field_count"]
    equiv_f = summary.get("globally_equiv_field_count", 0)
    mismatched_f = summary["mismatched_field_count"]
    total_f = summary["total_field_count"]
    matched_ratio = summary["matched_field_ratio"]
    mismatched_ratio = summary["mismatched_field_ratio"]
    diff_color = "red" if mismatched_f > 0 else "green"

    print()
    log_info("Value validation summary:")
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Matched transactions:"
        f" [bold green]{matched_tx}[/bold green]/[bold]{total_tx}[/bold]"
    )
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Matched fields (exact):"
        f" [bold green]{matched_f}[/bold green]/[bold]{total_f}[/bold]"
        f" ([green]{matched_ratio:.2f}%[/green])"
        + (f"  +  [bold cyan]{equiv_f}[/bold cyan] globally-equivalent" if equiv_f else "")
    )
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Unmatched fields:"
        f" [bold {diff_color}]{mismatched_f}[/bold {diff_color}]/[bold]{total_f}[/bold]"
        f" ([{diff_color}]{mismatched_ratio:.2f}%[/{diff_color}])"
    )


# ---------------------------------------------------------------------------
# Global equivalence groups
# ---------------------------------------------------------------------------
# Values within the same group are treated as semantically identical when
# the "global equivalence" mode is enabled during value validation.
_GLOBAL_EQUIV_GROUPS: list[frozenset[str]] = [
    # null-like vs zero/false/empty — e.g. \N vs 0, [null] vs 0, null vs 0
    frozenset({"\\n", "\\N", "", "null", "none", "0", "0.0", "false"}),
    # null-like vs empty array/object — e.g. [null] vs \N, [] vs null
    frozenset({"\\n", "\\N", "", "null", "none", "[]", "{}"}),
]

# Pre-computed lowercase versions for case-insensitive matching
_GLOBAL_EQUIV_GROUPS_LOWER: list[frozenset[str]] = [
    frozenset(v.lower() for v in grp) for grp in _GLOBAL_EQUIV_GROUPS
]

# Lowercase null-like sentinel set (used for recursive array normalization)
_NULL_LIKE_LOWER: frozenset[str] = frozenset({"\\n", "", "null", "none"})


def _normalize_for_global_equiv(value: Any) -> str:
    """Normalize *value* to a canonical lowercase string for equivalence checks.

    Arrays / tuples that contain **only** null-like elements are collapsed to
    ``"null"`` so that e.g. ``[None, None]`` matches ``None`` or ``"\\N"``.
    """
    if value is None:
        return "null"
    if isinstance(value, (list, tuple)):
        if len(value) == 0:
            return "[]"
        if all(_normalize_for_global_equiv(v) in _NULL_LIKE_LOWER for v in value):
            return "null"
        return str(value).lower()
    if isinstance(value, dict):
        return "{}" if len(value) == 0 else str(value).lower()
    return str(value).strip().lower()


def _globally_equivalent(a: Any, b: Any) -> bool:
    """Return ``True`` when *a* and *b* are considered globally equivalent.

    Equivalence is checked in three ways (in order):
    1. **Exact equality** via ``_null_safe_equal`` — identical values are always
       equivalent (short-circuits the heavier checks).
    2. **Group membership** — both values normalize to a string that belongs to
       the same ``_GLOBAL_EQUIV_GROUPS_LOWER`` group
       (e.g. ``None`` ↔ ``[]``, ``None`` ↔ ``0``).
    3. **Element-wise list comparison** — if both values are lists/tuples of the
       *same length*, each corresponding pair is recursively checked for
       equivalence.  This covers cases like ``[[], None]`` ≡ ``[None, None]``
       (SRC stores an empty list ``[]`` where BCV stores ``None`` for the same
       "no data" slot) without requiring the whole outer array to be null-like.
    """
    # 1. Exact / null-safe equality (fast path)
    if _null_safe_equal(a, b):
        return True
    na = _normalize_for_global_equiv(a)
    nb = _normalize_for_global_equiv(b)
    # 2. Direct group membership
    if any(na in grp and nb in grp for grp in _GLOBAL_EQUIV_GROUPS_LOWER):
        return True
    # 3. Element-wise comparison for same-length lists/tuples
    if (
        isinstance(a, (list, tuple))
        and isinstance(b, (list, tuple))
        and len(a) == len(b)
        and len(a) > 0
    ):
        return all(_globally_equivalent(ai, bi) for ai, bi in zip(a, b))
    return False


def _null_safe_equal(a: Any, b: Any) -> bool:
    """NULL-safe equality comparison.

    Handles the following cases that plain ``!=`` gets wrong:
    * ``None == None``  → True   (both SQL NULLs should match)
    * ``float('nan') == float('nan')`` → True  (IEEE-754 NaN ≠ NaN by default)
    """
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # IEEE-754 NaN: nan != nan is True, so treat both-NaN as equal.
    try:
        import math

        if isinstance(a, float) and isinstance(b, float) and math.isnan(a) and math.isnan(b):
            return True
    except Exception:
        pass
    return a == b


def _md_cell(value: Any) -> str:
    """Format a single markdown table cell value."""
    if value is None or value == "":
        return "*(null)*"
    return "`" + str(value).replace("`", "'").replace("|", "\\|") + "`"


def write_value_validation_report_md(
    table: str,
    summary: dict[str, Any],
    sql_sections: list[str] | None = None,
    now: datetime | None = None,
    parent_nodes: dict[str, str] | None = None,
    use_global_equiv: bool = False,
) -> Path:
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
    field_diffs: dict[str, list[dict[str, Any]]] = summary.get("field_diffs", {})
    mismatched_fields: list[str] = summary["mismatched_fields"]
    globally_equiv_fields: list[str] = summary.get("globally_equiv_fields", [])
    globally_equiv_diffs: dict[str, list[dict[str, Any]]] = summary.get("globally_equiv_diffs", {})
    equiv_f = summary.get("globally_equiv_field_count", 0)

    lines: list[str] = [
        f"# Value Validation Report — {table} — {timestamp}",
        "",
        "## Summary",
        "",
        "| Metric | Count | Ratio |",
        "|:---|---:|---:|",
        f"| SRC Transactions | {summary['total_transaction_count']} | |",
        f"| Matched Transactions (SRC ∩ BCV) | {summary['matched_transaction_count']}"
        f" | {summary['matched_transaction_count'] / summary['total_transaction_count'] * 100:.1f}%"
        f" |"
        if summary["total_transaction_count"]
        else f"| Matched Transactions (SRC ∩ BCV) | {summary['matched_transaction_count']} | |",
        f"| Total Fields | {summary['total_field_count']} | |",
        f"| Matched Fields (exact) | {summary['matched_field_count']}"
        f" | {summary['matched_field_ratio']:.1f}% |",
    ]
    if use_global_equiv:
        lines.append(
            f"| Globally Equivalent Fields | {equiv_f}"
            f" | {equiv_f / summary['total_field_count'] * 100:.1f}% |"
            if summary["total_field_count"]
            else f"| Globally Equivalent Fields | {equiv_f} | |"
        )
    lines += [
        f"| **Unmatched Fields** | **{summary['mismatched_field_count']}**"
        f" | **{summary['mismatched_field_ratio']:.1f}%** |",
        "",
    ]

    # ── Excluded Parent Structure Nodes ──────────────────────────────────────
    if parent_nodes:
        lines += [
            "## Excluded Parent Structure Nodes",
            "",
            f"The following **{len(parent_nodes)}** field(s) were skipped during value"
            " validation because they are parent structure nodes"
            " (type matches a structural type **and** at least one child column exists).",
            "",
            "| # | Field | Type |",
            "|---:|:---|:---|",
        ]
        for idx, (field, ftype) in enumerate(sorted(parent_nodes.items()), 1):
            lines.append(f"| {idx} | `{field}` | `{ftype}` |")
        lines.append("")

    # ── Globally Equivalent Fields ────────────────────────────────────────────
    if use_global_equiv and globally_equiv_fields:
        lines += [
            "## Globally Equivalent Fields",
            "",
            f"The following **{len(globally_equiv_fields)}** field(s) had value differences"
            " that are considered **semantically equivalent** under the global equivalence"
            " rules (e.g. `null` ↔ `0`, `[null]` ↔ `null`, `[]` ↔ `null`)."
            " They are counted as **matched** for the purposes of this report.",
            "",
            "| # | Field | Sample diffs |",
            "|---:|:---|:---|",
        ]
        for idx, field in enumerate(globally_equiv_fields, 1):
            diffs = globally_equiv_diffs.get(field, [])
            sample = diffs[0] if diffs else {}
            sample_text = f"`{sample.get('src')}` → `{sample.get('bcv')}`" if sample else ""
            lines.append(f"| {idx} | `{field}` | {sample_text} |")
        lines.append("")

    # ── SQL Queries ──────────────────────────────────────────────────────────
    if sql_sections:
        lines += ["## SQL Queries", ""]
        for section in sql_sections:
            lines += [section, ""]

    # ── Unmatched Field Details ───────────────────────────────────────────────
    if not field_diffs:
        lines += ["## Unmatched Field Details", "", "✅ No field mismatches found.", ""]
    else:
        lines += ["## Unmatched Field Details", ""]
        for idx, field in enumerate(mismatched_fields, 1):
            diffs = field_diffs.get(field, [])
            lines += [
                f"### {idx}. `{field}`  _({len(diffs)} diff(s))_",
                "",
                "| Key | SRC | BCV |",
                "|:---|:---|:---|",
            ]
            for diff in diffs:
                tid = _md_cell(diff["key"])
                src = _md_cell(diff["src"])
                bcv = _md_cell(diff["bcv"])
                lines.append(f"| {tid} | {src} | {bcv} |")
            lines.append("")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = OUTPUT_DIR / f"{table}_validation_report.md"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def update_result_csv_with_validation(
    result_path: Path,
    summary: dict[str, Any],
    parent_nodes: set[str],
) -> None:
    """Patch a 'validation' column (Y/E/N/-) into the existing result CSV.

    Values:
      Y  — exact match
      E  — globally equivalent (passes only when global-equiv mode is on)
      N  — mismatch
      -  — skipped (parent structure node)
    """
    matched_set = set(summary["matched_fields"])
    mismatched_set = set(summary["mismatched_fields"])
    globally_equiv_set = set(summary.get("globally_equiv_fields", []))

    with result_path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = list(reader.fieldnames or [])

    if "validation" not in fieldnames:
        fieldnames = fieldnames + ["validation"]

    for row in rows:
        if row.get("status") in ("MATCHED", "MATCHED - TYPE DIFF"):
            src_field = row.get("src_field", "")
            if src_field in parent_nodes:
                row["validation"] = "-"
            elif src_field in mismatched_set:
                row["validation"] = "N"
            elif src_field in globally_equiv_set:
                row["validation"] = "E"
            elif src_field in matched_set:
                row["validation"] = "Y"
            else:
                row["validation"] = ""
        else:
            row.setdefault("validation", "")

    with result_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def prompt_for_transaction_limit() -> int:
    selected = questionary.select(
        "Select number of transactions to sample for validation:",
        choices=[
            questionary.Choice(title="10", value=10),
            questionary.Choice(title="100", value=100),
            questionary.Choice(title="1000", value=1000),
        ],
    ).ask()
    return selected if selected is not None else 10


def prompt_for_value_validation() -> bool:
    selected = questionary.confirm("Continue to validate values for MATCHED columns?").ask()
    return bool(selected)


def prompt_for_global_equiv_mode() -> bool:
    """Ask whether semantically-equivalent values (null vs 0, [] vs null, …) should be treated as matching."""
    selected = questionary.confirm(
        "Enable global equivalence mode? "
        "(treats null/\\N/0/false/[]/[] as equivalent where applicable)"
    ).ask()
    return bool(selected)


def print_value_validation_sql(
    table: str,
    result_path: Path,
    src_connection_kwargs: dict[str, Any],
    bcv_connection_kwargs: dict[str, Any],
    now: datetime | None = None,
    batch_size: int = VALUE_VALIDATION_BATCH_SIZE,
) -> None:
    # Use a longer timeout for validation queries (they can be slow)
    src_connection_kwargs = {
        **src_connection_kwargs,
        "request_timeout": VALUE_VALIDATION_REQUEST_TIMEOUT,
    }
    bcv_connection_kwargs = {
        **bcv_connection_kwargs,
        "request_timeout": VALUE_VALIDATION_REQUEST_TIMEOUT,
    }

    matched_columns = load_matched_columns_from_result_csv(result_path)
    if not matched_columns:
        log_info("No MATCHED columns found for value validation.")
        return

    # Identify and skip parent structure nodes
    parent_nodes = identify_parent_structure_nodes(result_path)
    if parent_nodes:
        log_info(
            f"Skipping {len(parent_nodes)} parent structure node(s) from validation"
            f" (will be marked '-' in result CSV)"
        )
        matched_columns = [c for c in matched_columns if c not in parent_nodes]
    if not matched_columns:
        log_info("No columns left to validate after excluding parent structure nodes.")
        update_result_csv_with_validation(
            result_path,
            {
                "matched_fields": [],
                "mismatched_fields": [],
                "globally_equiv_fields": [],
            },
            parent_nodes,
        )
        return

    key_columns = get_table_key_columns(table)
    if not key_columns:
        log_info(f"No key columns configured for table '{table}'. Skipping value validation.")
        return

    transaction_limit = prompt_for_transaction_limit()
    log_info(f"Sampling up to {transaction_limit} transaction(s) for validation.")

    use_global_equiv = prompt_for_global_equiv_mode()
    if use_global_equiv:
        log_info("Global equivalence mode enabled: null/\\N/0/false/[]/[] treated as equivalent.")

    batch_id = build_value_validation_batch_id(now)
    sql_batches = build_value_validation_sql_batches(
        table,
        matched_columns,
        batch_id,
        key_columns,
        batch_size=batch_size,
        limit=transaction_limit,
    )
    num_batches = len(sql_batches)
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Value validation batch_id:"
        f" [bold green]{batch_id}[/bold green]"
        f"  (executing all {num_batches} batch(es))"
    )

    # ── Batch 1: TABLESAMPLE — anchors the key set ────────────────────────────
    batch0 = sql_batches[0]
    with spinning_cursor(f"Executing value validation batch 1/{num_batches} — SRC"):
        src_rows_0 = execute_sql(batch0["src_sql"], connection_kwargs=src_connection_kwargs)

    keys = get_row_keys(src_rows_0, key_columns)
    if not keys:
        log_info(
            f"No key values ({', '.join(key_columns)}) returned from SRC query. Skipping BCV query."
        )
        return

    bcv_sql_0 = build_value_validation_bcv_sql(
        table,
        batch0["columns"],
        batch_id,
        key_columns,
        keys,
        limit=transaction_limit,
    )
    with spinning_cursor(f"Executing value validation batch 1/{num_batches} — BCV"):
        bcv_rows_0 = execute_sql(bcv_sql_0, connection_kwargs=bcv_connection_kwargs)

    # Accumulators keyed by the composite row key tuple
    src_by_key: dict[tuple, dict[str, Any]] = {
        extract_row_key(r, key_columns): dict(r) for r in src_rows_0
    }
    bcv_by_key: dict[tuple, dict[str, Any]] = {
        extract_row_key(r, key_columns): dict(r) for r in bcv_rows_0
    }
    seen_columns: set[str] = set(batch0["columns"])
    all_columns: list[str] = list(batch0["columns"])
    sql_sections = [
        f"### Batch 1/{num_batches} — SRC (TABLESAMPLE)\n\n```sql\n{batch0['src_sql']}\n```",
        f"### Batch 1/{num_batches} — BCV\n\n```sql\n{bcv_sql_0}\n```",
    ]

    # ── Batch 2+: target the same keys, pull remaining columns ────────────────
    for batch_idx in range(1, num_batches):
        batch_num = batch_idx + 1
        batch = sql_batches[batch_idx]

        src_sql_n = build_value_validation_src_sql_by_keys(
            table,
            batch["columns"],
            batch_id,
            key_columns,
            keys,
        )
        with spinning_cursor(f"Executing value validation batch {batch_num}/{num_batches} — SRC"):
            src_rows_n = execute_sql(src_sql_n, connection_kwargs=src_connection_kwargs)

        bcv_sql_n = build_value_validation_bcv_sql(
            table,
            batch["columns"],
            batch_id,
            key_columns,
            keys,
            limit=transaction_limit,
        )
        with spinning_cursor(f"Executing value validation batch {batch_num}/{num_batches} — BCV"):
            bcv_rows_n = execute_sql(bcv_sql_n, connection_kwargs=bcv_connection_kwargs)

        for row in src_rows_n:
            k = extract_row_key(row, key_columns)
            if k in src_by_key:
                src_by_key[k].update(row)
        for row in bcv_rows_n:
            k = extract_row_key(row, key_columns)
            if k in bcv_by_key:
                bcv_by_key[k].update(row)
        for col in batch["columns"]:
            if col not in seen_columns:
                all_columns.append(col)
                seen_columns.add(col)
        sql_sections += [
            f"### Batch {batch_num}/{num_batches} — SRC\n\n```sql\n{src_sql_n}\n```",
            f"### Batch {batch_num}/{num_batches} — BCV\n\n```sql\n{bcv_sql_n}\n```",
        ]

    # ── Compare merged results ─────────────────────────────────────────────────
    summary = compare_value_validation_results(
        list(src_by_key.values()),
        list(bcv_by_key.values()),
        all_columns,
        key_columns,
        use_global_equiv=use_global_equiv,
    )
    print_value_validation_summary(summary)

    report_path = write_value_validation_report_md(
        table,
        summary,
        sql_sections=sql_sections,
        parent_nodes=parent_nodes,
        use_global_equiv=use_global_equiv,
    )
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Validation report written to: [bold green]{report_path}[/bold green]"
    )

    update_result_csv_with_validation(result_path, summary, parent_nodes)
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Result CSV updated with validation column: [bold green]{result_path}[/bold green]"
    )


def build_connection_kwargs(
    host: str,
    port: int,
    user: str,
    catalog: str,
    schema: str,
    request_timeout: float,
    auth_token: str | None,
    auth_header: str | None,
) -> dict[str, Any]:
    return {
        "host": host,
        "port": port,
        "user": user,
        "catalog": catalog,
        "schema": schema,
        "request_timeout": request_timeout,
        "http_headers": build_http_headers(auth_token, auth_header),
    }


def retrieve_column_list(
    table_name: str,
    describe_sql: str,
    connection_kwargs: dict[str, Any],
) -> list[dict[str, Any]]:
    log_info(f"Retrieving column list for table: {table_name}")
    with spinning_cursor():
        result = execute_sql(describe_sql, connection_kwargs=connection_kwargs)
    return result


def resolve_table(table: str | None) -> str:
    if table:
        if table not in ALLOWED_TABLES:
            raise SystemExit(f"Invalid table: {table}")
        return table

    UI_CONSOLE.print(
        Panel(
            f"{APP_BANNER}\nBCV Analyzer {APP_VERSION}\nUse ↑/↓ to choose a table, then press Enter.\n\n"
            "[bold]Analysis Summary Rules[/bold] (applied to DIFF rows where SRC has value, BCV is missing):\n"
            "  [bold cyan]■ Recommended for Backfill   [/bold cyan]: usage (ETL = [bold]Y[/bold] OR SOS = [bold]Y[/bold] OR Insights [bold]> 0[/bold] OR Arena [bold]> 0[/bold] OR LQS [bold]≥ 10[/bold] OR CP [bold]> 0[/bold] OR AF [bold]> 0[/bold] OR Others [bold]≥ 100[/bold])  AND  size < 0.03 TiB (or unknown)\n"
            "  [bold yellow]■ Recommended Excluded       [/bold yellow]: usage (ETL = [bold]Y[/bold] OR SOS = [bold]Y[/bold] OR Insights [bold]> 0[/bold] OR Arena [bold]> 0[/bold] OR LQS [bold]≥ 10[/bold] OR CP [bold]> 0[/bold] OR AF [bold]> 0[/bold] OR Others [bold]≥ 100[/bold])  AND  size ≥ 0.03 TiB\n"
            "  [bold red]■ Recommended No Backfill    [/bold red]: usage below threshold  (ETL is blank, SOS is blank, Insights = 0, Arena = 0, LQS < 10, CP = 0, AF = 0, Others < 100)",
            title="BCV Analyzer",
            border_style="cyan",
        )
    )
    selected = questionary.select("Select table:", choices=list(ALLOWED_TABLES)).ask()
    if not selected:
        raise SystemExit("No table selected")
    return selected


def resolve_run_mode() -> RunMode:
    selected = questionary.select(
        "Select run mode:",
        choices=[
            questionary.Choice(
                title="Full Run  (query Presto for ALL missing columns)",
                value=RunMode.FULL_RUN,
            ),
            questionary.Choice(
                title="Validation Only  (skip schema comparison, load existing result.csv and validate data values)",
                value=RunMode.VALIDATION_ONLY,
            ),
        ],
    ).ask()
    if not selected:
        raise SystemExit("No run mode selected")
    return selected


def validate_connection_args(host: str | None, user: str | None) -> None:
    required = {
        "host": host,
        "user": user,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise SystemExit(f"Missing required connection args: {', '.join(missing)}")


def get_usage_int(row: dict[str, Any], column: str) -> int:
    try:
        return int(row.get(column) or 0)
    except (ValueError, TypeError):
        return 0


def is_etl_used(row: dict[str, Any]) -> bool:
    return row.get("usage:ETL") == "Y"


def is_sos_used(row: dict[str, Any]) -> bool:
    return row.get("usage:SOS") == "Y"


def usage_meets_threshold(row: dict[str, Any]) -> bool:
    return (
        is_etl_used(row)
        or is_sos_used(row)
        or get_usage_int(row, "usage:Insights") > 0
        or get_usage_int(row, "usage:Arena") > 0
        or get_usage_int(row, "usage:LQS") >= 10
        or get_usage_int(row, "usage:CP") > 0
        or get_usage_int(row, "usage:AF") > 0
        or get_usage_int(row, "usage:Others") >= 100
    )


def get_recommended_action(row: dict[str, Any], was_queried: bool) -> str:
    if not is_missing_bcv_column(row):
        return ""
    if not was_queried and not is_etl_used(row) and not is_sos_used(row):
        return ""
    if usage_meets_threshold(row):
        size = row.get("size")
        if not isinstance(size, (int, float)) or float(size) < 0.03:
            return "Backfill"
        return "Excluded - Size Too Large"
    return "No Backfill - Low Usage"


def print_analysis_summary(
    queried_rows: list[dict[str, Any]],
    type_mismatch_rows: list[dict[str, Any]] | None = None,
    excluded_rows: list[dict[str, Any]] | None = None,
) -> None:
    from rich.table import Table

    UI_CONSOLE.print()

    # --- Grey: excluded columns (shown first) ---
    if excluded_rows:
        ex_table = Table(
            show_header=True, header_style="bold white", border_style="white", show_lines=True
        )
        ex_table.add_column("#", justify="right", style="dim", no_wrap=True)
        ex_table.add_column("Column Name", style="white", no_wrap=True)
        ex_table.add_column("Status", style="dim", no_wrap=True)
        ex_table.add_column("SRC Type", style="dim", no_wrap=True)
        ex_table.add_column("BCV Type", style="dim", no_wrap=True)
        for i, row in enumerate(excluded_rows, 1):
            ex_table.add_row(
                str(i),
                str(row.get("src_field", "")),
                str(row.get("status", "")),
                str(row.get("src_type", "")),
                str(row.get("bcv_type", "")),
            )
        UI_CONSOLE.print(
            Panel(
                ex_table,
                title=f"[bold white]Excluded Columns (skipped from analysis & validation) — {len(excluded_rows)} column(s)[/bold white]",
                border_style="white",
            )
        )

    # All queried_rows already satisfy is_missing_bcv_column
    recommended = [
        row
        for row in queried_rows
        if usage_meets_threshold(row)
        and (not isinstance(row.get("size"), (int, float)) or float(row["size"]) < 0.03)
    ]

    excluded = [
        row
        for row in queried_rows
        if usage_meets_threshold(row)
        and isinstance(row.get("size"), (int, float))
        and float(row["size"]) >= 0.03
    ]

    low_usage = [row for row in queried_rows if not usage_meets_threshold(row)]

    UI_CONSOLE.print()

    # --- Blue: recommended for backfill ---
    if recommended:
        rec_table = Table(
            show_header=True, header_style="bold cyan", border_style="cyan", show_lines=True
        )
        rec_table.add_column("#", justify="right", style="dim", no_wrap=True)
        rec_table.add_column("Column Name", style="bold white", no_wrap=True)
        rec_table.add_column("Size (TiB)", justify="right", style="green", no_wrap=True)
        for column in USAGE_COLUMNS:
            rec_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(recommended, 1):
            rec_table.add_row(
                str(i),
                str(row["src_field"]),
                str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(
            Panel(
                rec_table,
                title=f"[bold cyan]Recommended for Backfill — {len(recommended)} column(s)[/bold cyan]",
                border_style="cyan",
            )
        )
    else:
        UI_CONSOLE.print(
            Panel(
                "No columns recommended for backfill.",
                title="[bold cyan]Recommended for Backfill[/bold cyan]",
                border_style="cyan",
            )
        )

    # --- Yellow: usage meets threshold but size too large ---
    if excluded:
        exc_table = Table(
            show_header=True, header_style="bold yellow", border_style="yellow", show_lines=True
        )
        exc_table.add_column("#", justify="right", style="dim", no_wrap=True)
        exc_table.add_column("Column Name", style="white", no_wrap=True)
        exc_table.add_column("Size (TiB)", justify="right", style="yellow", no_wrap=True)
        for column in USAGE_COLUMNS:
            exc_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(excluded, 1):
            exc_table.add_row(
                str(i),
                str(row["src_field"]),
                str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(
            Panel(
                exc_table,
                title=f"[bold yellow]Recommended Excluded (size ≥ 0.03 TiB) — {len(excluded)} column(s)[/bold yellow]",
                border_style="yellow",
            )
        )

    # --- Red: usage below threshold, no backfill ---
    if low_usage:
        low_table = Table(
            show_header=True, header_style="bold red", border_style="red", show_lines=True
        )
        low_table.add_column("#", justify="right", style="dim", no_wrap=True)
        low_table.add_column("Column Name", style="white", no_wrap=True)
        low_table.add_column("Size (TiB)", justify="right", no_wrap=True)
        for column in USAGE_COLUMNS:
            low_table.add_column(column, justify="right", no_wrap=True)
        for i, row in enumerate(low_usage, 1):
            low_table.add_row(
                str(i),
                str(row["src_field"]),
                str(row["size"]),
                *(str(row.get(column) or "") for column in USAGE_COLUMNS),
            )
        UI_CONSOLE.print(
            Panel(
                low_table,
                title=f"[bold red]Recommended No Backfill (usage below threshold) — {len(low_usage)} column(s)[/bold red]",
                border_style="red",
            )
        )

    # --- Magenta: field name matched but type differs ---
    if type_mismatch_rows:
        tm_table = Table(
            show_header=True, header_style="bold magenta", border_style="magenta", show_lines=True
        )
        tm_table.add_column("#", justify="right", style="dim", no_wrap=True)
        tm_table.add_column("Column Name", style="white", no_wrap=True)
        tm_table.add_column("SRC Type", style="magenta", no_wrap=True)
        tm_table.add_column("BCV Type", style="magenta", no_wrap=True)
        for i, row in enumerate(type_mismatch_rows, 1):
            tm_table.add_row(
                str(i),
                str(row["src_field"]),
                str(row["src_type"]),
                str(row["bcv_type"]),
            )
        UI_CONSOLE.print(
            Panel(
                tm_table,
                title=f"[bold magenta]Type Mismatch (field exists in BCV but type differs) — {len(type_mismatch_rows)} column(s)[/bold magenta]",
                border_style="magenta",
            )
        )


def main(
    host: Annotated[str | None, typer.Option(help="Presto host")] = os.getenv("PRESTO_HOST"),
    port: Annotated[int, typer.Option(help="Presto port")] = int(os.getenv("PRESTO_PORT", "8080")),
    user: Annotated[str | None, typer.Option(help="Presto user")] = os.getenv("PRESTO_USER"),
    request_timeout: Annotated[
        float,
        typer.Option("--request-timeout", help="Request timeout in seconds"),
    ] = float(os.getenv("PRESTO_REQUEST_TIMEOUT", "5")),
    auth_token: Annotated[
        str | None,
        typer.Option("--auth-token", help="Presto gateway token"),
    ] = os.getenv("PRESTO_AUTH_TOKEN"),
    auth_header: Annotated[
        str | None,
        typer.Option("--auth-header", help="Auth header name"),
    ] = os.getenv("PRESTO_AUTH_HEADER"),
    table: Annotated[str | None, typer.Option("--table", help="BCV table name")] = None,
) -> None:
    selected_table = resolve_table(table)
    run_mode = resolve_run_mode()
    log_info(f"Run mode: {run_mode.value}")
    validate_connection_args(host, user)
    src_connection_kwargs = build_connection_kwargs(
        host,
        port,
        user,
        SRC_CATALOG,
        SRC_SCHEMA,
        request_timeout,
        auth_token,
        auth_header,
    )
    bcv_connection_kwargs = build_connection_kwargs(
        host,
        port,
        user,
        BCV_CATALOG,
        BCV_SCHEMA,
        request_timeout,
        auth_token,
        auth_header,
    )

    # ── Validation Only: skip schema comparison & usage analysis ──────────────
    if run_mode == RunMode.VALIDATION_ONLY:
        result_filename = f"{selected_table}_result.csv"
        result_path = (OUTPUT_DIR / result_filename).resolve()
        if not result_path.exists():
            raise SystemExit(
                f"Result file not found: {result_path}\n"
                "Please run a Full Run analysis first to generate the result CSV."
            )
        log_info(f"Loading existing result file: {result_path}")
        print_value_validation_sql(
            selected_table, result_path, src_connection_kwargs, bcv_connection_kwargs
        )
        return
    # ──────────────────────────────────────────────────────────────────────────

    src_table_name = build_full_table_name(selected_table)
    bcv_table_name = build_bcv_table_name(selected_table)
    rows = retrieve_column_list(
        src_table_name,
        f"DESCRIBE {src_table_name}",
        src_connection_kwargs,
    )
    bcv_rows = retrieve_column_list(
        bcv_table_name,
        f"DESCRIBE {bcv_table_name}",
        bcv_connection_kwargs,
    )
    src_output_rows = remove_describe_metadata_columns(rows)
    bcv_output_rows = remove_describe_metadata_columns(bcv_rows)
    log_info("Retrieving column size")
    field_sizes = load_field_sizes(selected_table, FIELD_SIZE_DIR)
    write_rows_as_json_file(src_output_rows, selected_table)
    write_rows_as_json_file(bcv_output_rows, f"bcv_{selected_table}")
    comparison_rows = compare_schema_rows(
        selected_table, src_output_rows, bcv_output_rows, field_sizes
    )
    # ── Exclude configured columns ────────────────────────────────────────────
    excluded_fields = load_excluded_fields(selected_table, EXCLUDE_FIELDS_PATH)
    excluded_rows = [row for row in comparison_rows if row.get("src_field") in excluded_fields]
    comparison_rows = [
        row for row in comparison_rows if row.get("src_field") not in excluded_fields
    ]
    if excluded_fields:
        log_info(f"Excluding {len(excluded_rows)} column(s) configured in exclude.csv")
    # ─────────────────────────────────────────────────────────────────────────
    add_etl_usage_info(selected_table, comparison_rows, ETL_FIELDS_PATH)
    add_sos_usage_info(selected_table, comparison_rows, SOS_FIELDS_PATH)
    comparison_rows, queried_rows = add_usage_info(
        selected_table, comparison_rows, src_connection_kwargs
    )
    queried_fields = {str(r["src_field"]) for r in queried_rows}
    summary_rows = list(queried_rows)
    summary_fields = set(queried_fields)
    for row in comparison_rows:
        src_field = str(row.get("src_field") or "")
        if (
            is_missing_bcv_column(row)
            and (is_etl_used(row) or is_sos_used(row))
            and src_field not in summary_fields
        ):
            summary_rows.append(row)
            summary_fields.add(src_field)
    for row in comparison_rows:
        was_queried = is_missing_bcv_column(row) and str(row.get("src_field")) in queried_fields
        row["recommended_action"] = get_recommended_action(row, was_queried)
    type_mismatch_rows = [row for row in comparison_rows if is_type_mismatch(row)]
    print_analysis_summary(
        summary_rows, type_mismatch_rows=type_mismatch_rows, excluded_rows=excluded_rows or None
    )
    result_filename = f"{selected_table}_result.csv"
    write_rows_as_csv_file(
        comparison_rows,
        result_filename,
        (
            "status",
            "src_field",
            "src_type",
            "bcv_field",
            "bcv_type",
            "size",
            *USAGE_COLUMNS,
            "recommended_action",
        ),
    )
    result_path = (OUTPUT_DIR / result_filename).resolve()
    UI_CONSOLE.print(
        f"[{current_timestamp()}] Completed! results are written to [bold green]{result_path}[/bold green]"
    )
    if prompt_for_value_validation():
        print_value_validation_sql(
            selected_table, result_path, src_connection_kwargs, bcv_connection_kwargs
        )


if __name__ == "__main__":
    typer.run(main)
